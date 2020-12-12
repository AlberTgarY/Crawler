import re
import openpyxl
from stopwords import stopword
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer


def list_3_ngram(news_content, n=3, m=2):
    gram_three = []
    stopwords = stopword()
    # cut words
    words_list = stopwords.Word_cut_list(news_content)
    print(words_list)
    if len(words_list) < n:
        n = len(words_list)
    temp = [words_list[i - k:i] for k in range(m, n + 1) for i in range(k, len(words_list) + 1) ]

    for gram in temp:
        valid = True
        for word in gram:
            # get rid of the numbers
            if re.match(u"([\u0030-\u0039a])", word) or len(word) > 20:
                valid = False
                break
        if valid:
             gram_three.append(gram)

    return gram_three


def tf_idf(news_content):
    temp = []
    gram_three = list_3_ngram(str(news_content), n=3, m=2)
    if not gram_three:
        return None
    temp.extend([' '.join(['_'.join(i) for i in gram_three])])

    vectorizer = CountVectorizer()  # count frequency
    transformer = TfidfTransformer()  # tf-tdf weight

    freq = vectorizer.fit_transform(temp)  # freq.toarray() frequency array
    tfidf = transformer.fit_transform(freq)  # tfidf.toarray() tfidf matrix

    # create a tuple of index and tdidf weight
    index_keyword = []
    for i, tfidf_i in enumerate(tfidf.toarray()):
        index_keyword.append([(j, value) for j, value in enumerate(tfidf_i)])

    # sort descending
    index_keyword = [sorted(i, key=lambda x: x[1], reverse=True) for i in index_keyword]
    # reverse list, locating index
    index_keyword = [[j[0] for j in i] for i in index_keyword]

    tfidf_dic = vectorizer.vocabulary_
    tfidf_dic = dict(zip(tfidf_dic.values(), tfidf_dic.keys()))
    list_keyword = []

    for i in index_keyword:
        list_keyword.append([tfidf_dic[j] for j in i])
    list_keyword = [i[:30] for i in list_keyword]
    return list_keyword


def read_csv():
    key_word_list = []
    wb = openpyxl.load_workbook('../temp/2020-12-11.xlsx')
    for sheet in wb.sheetnames:
        c = wb[str(sheet)]
        for row in range(2, c.max_row):
            news_content = c.cell(row=row, column=4).value
            print("Content: "+news_content+"\n")
            temp = tf_idf(news_content)
            if temp:
                key_word_list.append(temp)
    for l in key_word_list:
        print(l)
    print(len(key_word_list))

read_csv()